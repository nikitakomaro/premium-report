import streamlit as st
import pandas as pd
import io
import os
import re
import warnings
import tempfile
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from bidi.algorithm import get_display

warnings.filterwarnings('ignore')

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="דוח פרמיה | Surense",
    page_icon="📊",
    layout="centered",
)

# ── RTL CSS ───────────────────────────────────────────────────────────────────
st.markdown("""
<style>
    body, .stApp { direction: rtl; }
    .stFileUploader, .stDownloadButton, .stAlert, .stMetric,
    .stMarkdown, h1, h2, h3, p, label { text-align: right; direction: rtl; }
    .stDownloadButton button { width: 100%; font-size: 15px; padding: 10px; }
    .metric-box {
        background: #f0f4ff;
        border-radius: 10px;
        padding: 16px 20px;
        text-align: center;
        border: 1px solid #d0d8f0;
    }
    .metric-box .val { font-size: 2rem; font-weight: bold; color: #1F4E79; }
    .metric-box .lbl { font-size: 0.85rem; color: #555; margin-top: 4px; }
    .red   { color: #CC0000 !important; }
    .green { color: #1A5C1A !important; }
    .section-divider { border-top: 2px solid #e0e0e0; margin: 24px 0; }
</style>
""", unsafe_allow_html=True)

# ── Hebrew PDF helper ─────────────────────────────────────────────────────────
_font_registered = False
BASE_FONT = 'Helvetica'

def _register_font():
    global _font_registered, BASE_FONT
    if _font_registered:
        return
    script_dir = os.path.dirname(os.path.abspath(__file__))
    candidates = [
        os.path.join(script_dir, 'Alef-Regular.ttf'),
        '/System/Library/Fonts/Supplemental/Arial Unicode.ttf',
        '/Library/Fonts/Arial Unicode.ttf',
        '/System/Library/Fonts/Supplemental/Arial.ttf',
        '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
        '/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf',
        '/usr/share/fonts/truetype/freefont/FreeSans.ttf',
    ]
    for fp in candidates:
        if os.path.exists(fp):
            try:
                pdfmetrics.registerFont(TTFont('HebFont', fp))
                BASE_FONT = 'HebFont'
                break
            except:
                pass
    _font_registered = True

def rh(text):
    return get_display(str(text)) if text is not None else ''

# ── Core analysis ─────────────────────────────────────────────────────────────
COL_POLICY  = "מס' חשבון/פוליסה"
COL_ID      = 'מספר ת.ז'
COL_FNAME   = 'שם פרטי לקוח'
COL_LNAME   = 'שם משפחה לקוח'
COL_PREMIUM = 'סה"כ פרמיה'
COL_MFG     = 'יצרן'
COL_PRODUCT = 'מוצר'
COL_AGENT   = 'מת"ל'
SHEET       = 'מוצרי ביטוח'
THRESHOLD   = 15.0

thin = Side(style='thin', color='CCCCCC')
BORD = Border(left=thin, right=thin, top=thin, bottom=thin)
CTR  = Alignment(horizontal='center', vertical='center')
RGT  = Alignment(horizontal='right',  vertical='center')
DFNT = Font(name='Arial', size=10)

SAVINGS_TYPES  = ['קרן השתלמות', 'קופת גמל לתגמולים ופיצויים', 'קופת גמל להשקעה']
PENSION_TYPES  = ['קרן פנסיה חדשה מקיפה', 'קרן פנסיה חדשה כללית', 'קרן פנסיה ותיקה']
PENSION_DEPOSIT_THRESHOLD = 0.02   # 2% מהפקדה

def get_fee_threshold(total_savings):
    if total_savings > 1_000_000:
        return 0.0065
    elif total_savings > 500_000:
        return 0.007
    elif total_savings > 250_000:
        return 0.0075
    else:
        return 0.008

def get_fee_reason(total_savings):
    if total_savings > 1_000_000:
        return 'מעל ₪1M — מקס׳ 0.65%'
    elif total_savings > 500_000:
        return 'מעל ₪500K — מקס׳ 0.70%'
    elif total_savings > 250_000:
        return 'מעל ₪250K — מקס׳ 0.75%'
    else:
        return 'דמי ניהול מעל 0.80%'

def analyze_management_fees(file2_bytes):
    df_raw = pd.read_excel(io.BytesIO(file2_bytes), sheet_name='מוצרי חיסכון')
    df = df_raw[df_raw['סוג מוצר'].isin(SAVINGS_TYPES)].copy()
    df = df[df['סטטוס מוצר'] == 'פעיל'].copy()
    df['צבירה']               = pd.to_numeric(df['צבירה'], errors='coerce').fillna(0)
    df['דמי ניהול מצבירה']   = pd.to_numeric(df['דמי ניהול מצבירה'], errors='coerce')
    df[COL_ID]                 = df[COL_ID].astype(str).str.strip()

    # צבירה כוללת לפי לקוח — סכום כל 3 סוגי המוצרים יחד
    savings_per_customer = df.groupby(COL_ID)['צבירה'].sum().reset_index()
    savings_per_customer.columns = [COL_ID, 'צבירה כוללת']

    df = df.merge(savings_per_customer, on=COL_ID)
    df['סף מקסימלי'] = df['צבירה כוללת'].apply(get_fee_threshold)
    df['סיבת חריגה'] = df['צבירה כוללת'].apply(get_fee_reason)

    # שורה נפרדת לכל מוצר עם חריגה
    exc = df[
        df['דמי ניהול מצבירה'].notna() &
        (df['דמי ניהול מצבירה'] > df['סף מקסימלי'])
    ].copy()
    exc['שם לקוח'] = exc['שם פרטי לקוח'].fillna('') + ' ' + exc['שם משפחה לקוח'].fillna('')
    exc['סוג מוצר'] = exc['סוג מוצר'].replace('קופת גמל לתגמולים ופיצויים', 'קופת גמל')
    exc = exc.sort_values(['צבירה כוללת', 'דמי ניהול מצבירה'], ascending=[False, False]).reset_index(drop=True)

    # פירוט צבירה לפי סוג מוצר לכל לקוח חריג (לdebug)
    exc_ids = exc[COL_ID].unique().tolist()
    breakdown = df[df[COL_ID].isin(exc_ids)].groupby([COL_ID, 'סוג מוצר'])['צבירה'].sum().reset_index()
    breakdown.columns = [COL_ID, 'סוג מוצר', 'צבירה']

    return exc, breakdown

def analyze_pension_fees(file2_bytes):
    """חריגות דמי ניהול בקרנות פנסיה — מהפקדה ומצבירה."""
    df_raw = pd.read_excel(io.BytesIO(file2_bytes), sheet_name='מוצרי חיסכון')
    df = df_raw[df_raw['סוג מוצר'].isin(PENSION_TYPES)].copy()
    df = df[df['סטטוס מוצר'] == 'פעיל'].copy()
    df['צבירה']                  = pd.to_numeric(df['צבירה'],                  errors='coerce').fillna(0)
    df['דמי ניהול מהפקדה']      = pd.to_numeric(df['דמי ניהול מהפקדה'],      errors='coerce')
    df['דמי ניהול מצבירה']      = pd.to_numeric(df['דמי ניהול מצבירה'],      errors='coerce')
    df[COL_ID]                    = df[COL_ID].astype(str).str.strip()
    df['שם לקוח']                = df['שם פרטי לקוח'].fillna('') + ' ' + df['שם משפחה לקוח'].fillna('')

    # קיצור שמות סוג מוצר
    PENSION_SHORT = {
        'קרן פנסיה חדשה מקיפה': 'פנסיה מקיפה',
        'קרן פנסיה חדשה כללית': 'פנסיה כללית',
        'קרן פנסיה ותיקה':       'פנסיה ותיקה',
    }
    df['סוג מוצר'] = df['סוג מוצר'].replace(PENSION_SHORT)

    # שם החברה — עמודת יצרן
    if 'יצרן' not in df.columns and 'שם יצרן' in df.columns:
        df['יצרן'] = df['שם יצרן']

    # קיצור שם החברה — הסרת "מקפת קרנות פנסיה וגמל", "פנסיה וגמל בע״מ" וכיו"ב
    def _shorten_company(name):
        if pd.isna(name):
            return name
        s = str(name)
        # הסר הכל מ-"מקפת" ואילך (מגדל מקפת קרנות פנסיה וגמל → מגדל)
        s = re.sub(r'\s*מקפת\b.*', '', s)
        # הסר הכל מ-"פנסיה וגמל" ואילך
        s = re.sub(r'\s*פנסיה\s+וגמל\b.*', '', s)
        # הסר "בע״מ" שיישאר בסוף
        s = re.sub(r'\s*בע["\u05f4\u2019]?מ\.?$', '', s)
        return s.strip()
    if 'יצרן' in df.columns:
        df['יצרן'] = df['יצרן'].apply(_shorten_company)

    # ── טבלה 1: חריגות דמי ניהול מהפקדה (>2%) ──
    exc1 = df[
        df['דמי ניהול מהפקדה'].notna() &
        (df['דמי ניהול מהפקדה'] > PENSION_DEPOSIT_THRESHOLD)
    ].copy()
    exc1 = exc1.sort_values('דמי ניהול מהפקדה', ascending=False).reset_index(drop=True)

    # ── טבלה 2: חריגות דמי ניהול מצבירה (לפי מדרגות) ──
    def pension_acc_threshold(s):
        if s > 1_000_000: return 0.0005
        if s > 500_000:   return 0.0008
        if s > 200_000:   return 0.001
        return None  # מתחת 200K — לא נבדק

    def pension_acc_reason(s):
        if s > 1_000_000: return 'מעל ₪1M — מקס׳ 0.05%'
        if s > 500_000:   return 'מעל ₪500K — מקס׳ 0.08%'
        if s > 200_000:   return 'מעל ₪200K — מקס׳ 0.10%'
        return None

    # צבירה כוללת לפי לקוח (בתוך קרנות פנסיה בלבד)
    totals = df.groupby(COL_ID)['צבירה'].sum().reset_index()
    totals.columns = [COL_ID, 'צבירה כוללת']
    df2 = df.merge(totals, on=COL_ID)
    df2['סף מצבירה']   = df2['צבירה כוללת'].apply(pension_acc_threshold)
    df2['סיבת חריגה']  = df2['צבירה כוללת'].apply(pension_acc_reason)

    exc2 = df2[
        df2['דמי ניהול מצבירה'].notna() &
        df2['סף מצבירה'].notna() &
        (df2['דמי ניהול מצבירה'] > df2['סף מצבירה'])
    ].copy()
    exc2.rename(columns={'סף מצבירה': 'סף מקסימלי'}, inplace=True)
    exc2 = exc2.sort_values(['צבירה כוללת', 'דמי ניהול מצבירה'], ascending=[False, False]).reset_index(drop=True)

    return exc1, exc2

def analyze_gold_customers(file2_bytes):
    """לקוחות עם צבירה כוללת מעל מיליון ש״ח במוצרי חיסכון."""
    df_raw = pd.read_excel(io.BytesIO(file2_bytes), sheet_name='מוצרי חיסכון')
    df = df_raw[df_raw['סוג מוצר'].isin(SAVINGS_TYPES)].copy()
    df = df[df['סטטוס מוצר'] == 'פעיל'].copy()
    df['צבירה'] = pd.to_numeric(df['צבירה'], errors='coerce').fillna(0)
    df[COL_ID]  = df[COL_ID].astype(str).str.strip()

    # צבירה כוללת לפי לקוח
    totals = df.groupby(COL_ID)['צבירה'].sum().reset_index()
    totals.columns = [COL_ID, 'צבירה כוללת']

    # פירוט לפי סוג מוצר לכל לקוח
    by_type = df.groupby([COL_ID, 'סוג מוצר'])['צבירה'].sum().unstack(fill_value=0).reset_index()

    # דמי ניהול ממוצע משוקלל לפי לקוח
    df['דמי ניהול מצבירה'] = pd.to_numeric(df['דמי ניהול מצבירה'], errors='coerce')
    fee_df = df.dropna(subset=['דמי ניהול מצבירה']).copy()
    if len(fee_df) > 0:
        fee_df['weighted'] = fee_df['דמי ניהול מצבירה'] * fee_df['צבירה']
        fee_agg = fee_df.groupby(COL_ID).agg(weighted=('weighted','sum'), total=('צבירה','sum')).reset_index()
        fee_agg['דמי ניהול ממוצע'] = fee_agg['weighted'] / fee_agg['total'].replace(0, float('nan'))
        fee_agg = fee_agg[[COL_ID, 'דמי ניהול ממוצע']]
    else:
        fee_agg = pd.DataFrame(columns=[COL_ID, 'דמי ניהול ממוצע'])

    # פרטי לקוח (שם, סוכן) — שורה ראשונה לכל לקוח
    details = df.sort_values('צבירה', ascending=False).groupby(COL_ID, sort=False).agg(
        שם_פרטי  = ('שם פרטי לקוח',  'first'),
        שם_משפחה = ('שם משפחה לקוח', 'first'),
        agent     = (COL_AGENT,        'first'),
    ).reset_index()
    details['שם לקוח'] = details['שם_פרטי'].fillna('') + ' ' + details['שם_משפחה'].fillna('')

    result = totals.merge(details[[COL_ID, 'שם לקוח', 'agent']], on=COL_ID)
    result = result.merge(by_type, on=COL_ID, how='left')
    result = result.merge(fee_agg, on=COL_ID, how='left')

    # רק מעל מיליון
    result = result[result['צבירה כוללת'] > 1_000_000].copy()
    result.columns = [c.replace('קופת גמל לתגמולים ופיצויים', 'קופת גמל') for c in result.columns]
    result = result.sort_values('צבירה כוללת', ascending=False).reset_index(drop=True)
    result.rename(columns={'agent': COL_AGENT}, inplace=True)
    return result

def analyze(file1_bytes, file2_bytes):
    df1 = pd.read_excel(io.BytesIO(file1_bytes), sheet_name=SHEET)
    df2 = pd.read_excel(io.BytesIO(file2_bytes), sheet_name=SHEET)

    for df in [df1, df2]:
        df[COL_PREMIUM] = pd.to_numeric(df[COL_PREMIUM], errors='coerce')
        df[COL_POLICY]  = df[COL_POLICY].astype(str).str.strip()
        df[COL_ID]      = df[COL_ID].astype(str).str.strip()

    df1d = df1.sort_values(COL_PREMIUM, ascending=False).drop_duplicates(COL_POLICY)
    df2d = df2.sort_values(COL_PREMIUM, ascending=False).drop_duplicates(COL_POLICY)

    keep = [COL_POLICY, COL_ID, COL_FNAME, COL_LNAME, COL_MFG, COL_PRODUCT, COL_PREMIUM, COL_AGENT]
    merged = df1d[keep].merge(df2d[keep], on=COL_POLICY, suffixes=('_jan', '_feb'))
    merged['שם לקוח']      = merged[COL_FNAME+'_feb'].fillna(merged[COL_FNAME+'_jan']) + ' ' + merged[COL_LNAME+'_feb'].fillna(merged[COL_LNAME+'_jan'])
    merged['ת.ז']           = merged[COL_ID+'_feb'].fillna(merged[COL_ID+'_jan'])
    merged['יצרן']          = merged[COL_MFG+'_feb'].fillna(merged[COL_MFG+'_jan'])
    merged['מוצר']          = merged[COL_PRODUCT+'_feb'].fillna(merged[COL_PRODUCT+'_jan'])
    merged['מת"ל']          = merged[COL_AGENT+'_feb'].fillna(merged[COL_AGENT+'_jan'])
    merged['פרמיה קודמת']  = merged[COL_PREMIUM+'_jan']
    merged['פרמיה נוכחית'] = merged[COL_PREMIUM+'_feb']
    merged['עלייה ₪']       = merged['פרמיה נוכחית'] - merged['פרמיה קודמת']
    merged['עלייה %']       = (merged['עלייה ₪'] / merged['פרמיה קודמת'].replace(0, float('nan'))) * 100

    result = merged[merged['עלייה %'] > THRESHOLD].sort_values('עלייה %', ascending=False).reset_index(drop=True)

    gone_pol = set(df1d[COL_POLICY]) - set(df2d[COL_POLICY])
    new_pol  = set(df2d[COL_POLICY]) - set(df1d[COL_POLICY])
    gone_df  = df1d[df1d[COL_POLICY].isin(gone_pol)]
    new_df   = df2d[df2d[COL_POLICY].isin(new_pol)]

    return merged, result, gone_df, new_df, df1d, df2d

# ── Excel builder ─────────────────────────────────────────────────────────────
def build_excel(merged, result, gone_df, new_df, fee_exceptions=None, agent=None, gold_customers=None, pension_deposit_exc=None, pension_acc_exc=None):
    wb = Workbook()
    HDR = PatternFill('solid', start_color='1F4E79')
    HF  = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    RED = PatternFill('solid', start_color='FFCCCC')
    YEL = PatternFill('solid', start_color='FFF2CC')
    ALT = PatternFill('solid', start_color='EBF3FB')

    label = f' — {agent}' if agent else ''

    def hdr_row(ws, hdrs, wids, color='1F4E79'):
        fill = PatternFill('solid', start_color=color)
        for ci, (h, w) in enumerate(zip(hdrs, wids), 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font = HF; c.fill = fill; c.alignment = CTR; c.border = BORD
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f'A1:{get_column_letter(len(hdrs))}1'
        ws.sheet_view.rightToLeft = True

    # Sheet 1 — increase
    ws1 = wb.active
    ws1.title = 'עלייה בפרמיה >15%'
    hdrs1 = ['ת.ז','שם לקוח',"מס' פוליסה",'יצרן','מוצר','מת"ל','פרמיה קודמת (₪)','פרמיה נוכחית (₪)','עלייה (₪)','עלייה (%)']
    wids1 = [14,24,14,26,26,18,18,20,14,12]
    fmts1 = [None,None,None,None,None,None,'#,##0.00','#,##0.00','#,##0.00','0.0%']
    hdr_row(ws1, hdrs1, wids1)
    df_src = result[result['מת"ל'] == agent] if agent else result
    for ri, (_, row) in enumerate(df_src.iterrows()):
        pct  = row['עלייה %']
        fill = RED if pct > 30 else (YEL if ri % 2 == 0 else None)
        vals = [row['ת.ז'], row['שם לקוח'], row[COL_POLICY], row['יצרן'], row['מוצר'], row['מת"ל'],
                row['פרמיה קודמת'], row['פרמיה נוכחית'], row['עלייה ₪'], pct/100]
        for ci, (val, fmt) in enumerate(zip(vals, fmts1), 1):
            c = ws1.cell(row=ri+2, column=ci, value=val)
            c.font = DFNT; c.border = BORD
            c.alignment = CTR if ci in [1,3,10] else RGT
            if fill: c.fill = fill
            if fmt:  c.number_format = fmt

    # Sheet 2 — summary
    ws2 = wb.create_sheet('סיכום')
    ws2.sheet_view.rightToLeft = True
    ws2.column_dimensions['A'].width = 36
    ws2.column_dimensions['B'].width = 24
    TF = Font(name='Arial', bold=True, size=14, color='1F4E79')
    SF = Font(name='Arial', bold=True, size=11, color='1F4E79')

    t = ws2.cell(row=1, column=1, value=f'דוח עלייה בפרמיה{label}')
    t.font = TF; t.alignment = RGT; ws2.merge_cells('A1:B1')
    ws2.cell(row=2, column=1, value=f'הופק: {datetime.now().strftime("%d/%m/%Y")}').font = Font(name='Arial', size=10, color='666666')
    ws2.cell(row=2, column=1).alignment = RGT

    agent_data = merged[merged['מת"ל'] == agent] if agent else merged
    agent_result = result[result['מת"ל'] == agent] if agent else result

    def sr(ws, r, lbl, val, fmt=None, bold_v=False):
        lc = ws.cell(row=r, column=1, value=lbl)
        lc.font = Font(name='Arial', bold=True, size=11); lc.alignment = RGT
        vc = ws.cell(row=r, column=2, value=val)
        vc.font = Font(name='Arial', bold=bold_v, size=11); vc.alignment = RGT
        if fmt: vc.number_format = fmt
        return r + 1

    r = 4
    ws2.cell(row=r, column=1, value='נתוני פוליסות').font = SF; r += 1
    r = sr(ws2, r, 'סה"כ פוליסות שהושוו', len(agent_data))
    r = sr(ws2, r, 'פוליסות עם עלייה >15%', len(agent_result), bold_v=True)
    if len(agent_result) > 0:
        r = sr(ws2, r, 'ממוצע עלייה', agent_result['עלייה %'].mean()/100, '0.0%')
        r = sr(ws2, r, 'עלייה מקסימלית', agent_result['עלייה %'].max()/100, '0.0%')
        r = sr(ws2, r, 'סך עלייה חודשית (₪)', agent_result['עלייה ₪'].sum(), '#,##0.00', True)
    r += 1
    ws2.cell(row=r, column=1, value='פוליסות שנסגרו / חדשות').font = SF; r += 1
    g = gone_df[gone_df[COL_AGENT] == agent] if agent else gone_df
    n = new_df[new_df[COL_AGENT] == agent] if agent else new_df
    r = sr(ws2, r, 'פוליסות שנסגרו', len(g))
    r = sr(ws2, r, 'פוליסות חדשות', len(n))

    # Sheet 3 & 4 — gone / new
    def policy_sheet(ws, title, df_sub, color):
        ws.sheet_view.rightToLeft = True
        t = ws.cell(row=1, column=1, value=title)
        t.font = Font(name='Arial', bold=True, size=12, color=color); t.alignment = RGT
        ws.merge_cells(f'A1:{get_column_letter(7)}1')
        hdrs = ['ת.ז','שם פרטי','שם משפחה',"מס' פוליסה",'יצרן','מוצר','פרמיה (₪)']
        wids = [14,14,14,14,28,28,14]
        for ci, (h, w) in enumerate(zip(hdrs, wids), 1):
            c = ws.cell(row=2, column=ci, value=h)
            c.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
            c.fill = PatternFill('solid', start_color=color)
            c.alignment = CTR; c.border = BORD
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.freeze_panes = 'A3'
        alt = PatternFill('solid', start_color='FFE8E8' if color == '7B0000' else 'E8F5E8')
        for ri, (_, row) in enumerate(df_sub.iterrows()):
            fill = alt if ri % 2 == 0 else None
            vals = [row.get(COL_ID,''), row.get(COL_FNAME,''), row.get(COL_LNAME,''),
                    row.get(COL_POLICY,''), row.get(COL_MFG,''), row.get(COL_PRODUCT,''), row.get(COL_PREMIUM,0)]
            for ci, (val, fmt) in enumerate(zip(vals, [None]*6+['#,##0.00']), 1):
                c = ws.cell(row=ri+3, column=ci, value=val)
                c.font = DFNT; c.border = BORD; c.alignment = CTR if ci<=4 else RGT
                if fill: c.fill = fill
                if fmt:  c.number_format = fmt

    ws3 = wb.create_sheet('פוליסות שנסגרו')
    policy_sheet(ws3, f'פוליסות שנסגרו{label}', g, '7B0000')
    ws4 = wb.create_sheet('פוליסות חדשות')
    policy_sheet(ws4, f'פוליסות חדשות{label}', n, '1A5C1A')

    # Sheet 5 — management fee exceptions (only in combined report)
    if fee_exceptions is not None and len(fee_exceptions) > 0 and agent is None:
        ws5 = wb.create_sheet('חריגות דמי ניהול')
        ws5.sheet_view.rightToLeft = True
        hdrs5 = ['ת.ז','שם לקוח','סוג מוצר','מוצר','יצרן','צבירה כוללת (₪)','צבירה מוצר (₪)','דמי ניהול בפועל','סף מקסימלי','חריגה','מת"ל','סיבת חריגה']
        wids5 = [14,22,18,26,22,16,16,16,14,12,18,28]
        ORANGE = PatternFill('solid', start_color='FF6600')
        for ci, (h, w) in enumerate(zip(hdrs5, wids5), 1):
            c = ws5.cell(row=1, column=ci, value=h)
            c.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
            c.fill = ORANGE; c.alignment = CTR; c.border = BORD
            ws5.column_dimensions[get_column_letter(ci)].width = w
        ws5.row_dimensions[1].height = 22
        ws5.freeze_panes = 'A2'
        ws5.auto_filter.ref = f'A1:{get_column_letter(len(hdrs5))}1'
        LORG = PatternFill('solid', start_color='FFF0E0')
        for ri, (_, row) in enumerate(fee_exceptions.iterrows()):
            fee    = row.get('דמי ניהול מצבירה', 0) or 0
            thresh = row.get('סף מקסימלי', 0) or 0
            excess = (fee - thresh) * 100
            fill   = LORG if ri % 2 == 0 else None
            vals   = [row.get(COL_ID,''), row.get('שם לקוח',''),
                      row.get('סוג מוצר',''), row.get('מוצר',''), row.get('יצרן',''),
                      row.get('צבירה כוללת',0), row.get('צבירה',0),
                      fee, thresh, excess/100, row.get('מת"ל',''), row.get('סיבת חריגה','')]
            fmts   = [None,None,None,None,None,'#,##0','#,##0','0.000%','0.000%','0.000%',None,None]
            for ci, (val, fmt) in enumerate(zip(vals, fmts), 1):
                c = ws5.cell(row=ri+2, column=ci, value=val)
                c.font = DFNT; c.border = BORD
                c.alignment = CTR if ci in [1,8,9,10] else RGT
                if fill: c.fill = fill
                if fmt:  c.number_format = fmt

    # Sheet 6 — לקוחות זהב (only in combined report)
    if gold_customers is not None and len(gold_customers) > 0 and agent is None:
        ws6 = wb.create_sheet('לקוחות זהב')
        ws6.sheet_view.rightToLeft = True
        gold_type_cols = [c for c in ['קרן השתלמות', 'קופת גמל', 'קופת גמל להשקעה'] if c in gold_customers.columns]
        has_fee_col = 'דמי ניהול ממוצע' in gold_customers.columns
        hdrs6 = ['ת.ז', 'שם לקוח', 'מת"ל', 'צבירה כוללת (₪)'] + [f'{c} (₪)' for c in gold_type_cols]
        if has_fee_col: hdrs6.append('דמי ניהול ממוצע')
        wids6 = [14, 24, 18, 18] + [18]*len(gold_type_cols) + ([14] if has_fee_col else [])
        GOLD_H = PatternFill('solid', start_color='FFD700')
        GOLD_R = PatternFill('solid', start_color='FFFACD')
        for ci, (h, w) in enumerate(zip(hdrs6, wids6), 1):
            c = ws6.cell(row=1, column=ci, value=h)
            c.font = Font(name='Arial', bold=True, color='000000', size=10)
            c.fill = GOLD_H; c.alignment = CTR; c.border = BORD
            ws6.column_dimensions[get_column_letter(ci)].width = w
        ws6.row_dimensions[1].height = 22
        ws6.freeze_panes = 'A2'
        ws6.auto_filter.ref = f'A1:{get_column_letter(len(hdrs6))}1'
        for ri, (_, row) in enumerate(gold_customers.iterrows()):
            fill = GOLD_R if ri % 2 == 0 else None
            vals = [row.get(COL_ID,''), row.get('שם לקוח',''), row.get(COL_AGENT,''),
                    row.get('צבירה כוללת', 0)] + [row.get(c, 0) for c in gold_type_cols]
            if has_fee_col: vals.append(row.get('דמי ניהול ממוצע', None))
            fmts = [None, None, None] + ['#,##0'] * (len(gold_type_cols) + 1) + (['0.000%'] if has_fee_col else [])
            for ci, (val, fmt) in enumerate(zip(vals, fmts), 1):
                c = ws6.cell(row=ri+2, column=ci, value=val)
                c.font = DFNT; c.border = BORD
                c.alignment = RGT
                if fill: c.fill = fill
                if fmt:  c.number_format = fmt

    # Sheet 7 — חריגות פנסיה מהפקדה
    def pension_sheet(ws, title, df_p, hdr_color, cols, col_fmts):
        ws.sheet_view.rightToLeft = True
        PHDR = PatternFill('solid', start_color=hdr_color)
        PALE = PatternFill('solid', start_color='F0E6FF')
        wids = [14,22,18,22,14,14,14,14,28]
        for ci, (h, w) in enumerate(zip(cols, wids[:len(cols)]), 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
            c.fill = PHDR; c.alignment = CTR; c.border = BORD
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f'A1:{get_column_letter(len(cols))}1'
        for ri, (_, row) in enumerate(df_p.iterrows()):
            fill = PALE if ri % 2 == 0 else None
            vals = [row.get(c, '') for c in col_fmts.keys()]
            fmts = list(col_fmts.values())
            for ci, (val, fmt) in enumerate(zip(vals, fmts), 1):
                c = ws.cell(row=ri+2, column=ci, value=val)
                c.font = DFNT; c.border = BORD; c.alignment = RGT
                if fill: c.fill = fill
                if fmt:  c.number_format = fmt

    if pension_deposit_exc is not None and len(pension_deposit_exc) > 0 and agent is None:
        ws7 = wb.create_sheet('פנסיה — חריגות הפקדה')
        p1_col_fmts = {COL_ID: None, 'שם לקוח': None, COL_AGENT: None,
                       'סוג מוצר': None, 'מוצר': None, 'צבירה': '#,##0',
                       'דמי ניהול מהפקדה': '0.000%'}
        pension_sheet(ws7, 'חריגות דמי ניהול מהפקדה — קרן פנסיה',
                      pension_deposit_exc, '6A0DAD', list(p1_col_fmts.keys()), p1_col_fmts)

    if pension_acc_exc is not None and len(pension_acc_exc) > 0 and agent is None:
        ws8 = wb.create_sheet('פנסיה — חריגות צבירה')
        p2_col_fmts = {COL_ID: None, 'שם לקוח': None, COL_AGENT: None,
                       'סוג מוצר': None, 'מוצר': None, 'צבירה': '#,##0',
                       'צבירה כוללת': '#,##0', 'דמי ניהול מצבירה': '0.000%',
                       'סף מקסימלי': '0.000%', 'סיבת חריגה': None}
        pension_sheet(ws8, 'חריגות דמי ניהול מצבירה — קרן פנסיה',
                      pension_acc_exc, '154360', list(p2_col_fmts.keys()), p2_col_fmts)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ── PDF builder ───────────────────────────────────────────────────────────────
def build_pdf(merged, result, gone_df, new_df, month_label, agent=None, fee_exceptions=None, gold_customers=None, pension_deposit_exc=None, pension_acc_exc=None):
    _register_font()
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            rightMargin=1.5*cm, leftMargin=1.5*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    title_s = ParagraphStyle('t', fontName=BASE_FONT, fontSize=16,
                              textColor=colors.HexColor('#1F4E79'), spaceAfter=4, alignment=2)
    sub_s   = ParagraphStyle('s', fontName=BASE_FONT, fontSize=10,
                              textColor=colors.grey, spaceAfter=8, alignment=2)
    sec_s   = ParagraphStyle('h', fontName=BASE_FONT, fontSize=12,
                              textColor=colors.HexColor('#1F4E79'), spaceBefore=10, spaceAfter=6, alignment=2)

    agent_data   = merged[merged['מת"ל'] == agent] if agent else merged
    agent_result = result[result['מת"ל'] == agent] if agent else result
    g = gone_df[gone_df[COL_AGENT] == agent] if agent else gone_df
    n = new_df[new_df[COL_AGENT] == agent] if agent else new_df

    def page_header(ttl):
        """Returns title + subtitle paragraphs for the top of each page."""
        items = [Paragraph(rh(ttl), title_s)]
        if agent:
            items.append(Paragraph(rh(f'סוכן: {agent}'), sub_s))
        items.append(Paragraph(rh(month_label), sub_s))
        items.append(Paragraph(rh(f'הופק: {datetime.now().strftime("%d/%m/%Y")}'), sub_s))
        items.append(Spacer(1, 0.3*cm))
        return items

    story = []

    # ══════════════════════════════════════════════
    # עמוד 1 — סיכום + חריגות פרמיה
    # ══════════════════════════════════════════════
    story += page_header('דוח עלייה בפרמיה חודשית — סיכום וחריגות')

    story.append(Paragraph(rh('סיכום'), sec_s))
    sum_data = [[rh('נושא'), rh('ערך')],
                [rh('פוליסות שהושוו'), str(len(agent_data))],
                [rh('פוליסות עם עלייה >15%'), str(len(agent_result))]]
    if len(agent_result) > 0:
        sum_data += [
            [rh('ממוצע עלייה'), f"{agent_result['עלייה %'].mean():.1f}%"],
            [rh('עלייה מקסימלית'), f"{agent_result['עלייה %'].max():.1f}%"],
            [rh('סך עלייה חודשית'), f"₪{agent_result['עלייה ₪'].sum():,.0f}"],
        ]
    sum_data += [[rh('פוליסות שנסגרו'), str(len(g))],
                 [rh('פוליסות חדשות'), str(len(n))]]

    st_tbl = Table(sum_data, colWidths=[10*cm, 6*cm])
    st_tbl.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#1F4E79')),
        ('TEXTCOLOR', (0,0),(-1,0),colors.white),
        ('FONTNAME',  (0,0),(-1,-1),BASE_FONT),
        ('FONTSIZE',  (0,0),(-1,0),11),('FONTSIZE',(0,1),(-1,-1),10),
        ('ALIGN',     (0,0),(-1,-1),'RIGHT'),
        ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,colors.HexColor('#EBF3FB')]),
        ('GRID',      (0,0),(-1,-1),0.5,colors.HexColor('#CCCCCC')),
        ('TOPPADDING',(0,0),(-1,-1),5),('BOTTOMPADDING',(0,0),(-1,-1),5),
    ]))
    story.append(st_tbl)
    story.append(Spacer(1, 0.5*cm))

    if len(agent_result) > 0:
        story.append(Paragraph(rh(f'פוליסות עם עלייה >15% ({len(agent_result)})'), sec_s))
        th = [rh('ת.ז'),rh('שם לקוח'),rh("מס' פוליסה"),rh('יצרן'),
              rh('פרמיה קודמת'),rh('פרמיה נוכחית'),rh('עלייה ₪'),rh('עלייה %')]
        td = [th]
        for _, row in agent_result.iterrows():
            td.append([rh(row['ת.ז']),rh(row['שם לקוח']),rh(row[COL_POLICY]),rh(row['יצרן']),
                       f"₪{row['פרמיה קודמת']:,.0f}",f"₪{row['פרמיה נוכחית']:,.0f}",
                       f"₪{row['עלייה ₪']:,.0f}",f"{row['עלייה %']:.1f}%"])
        mt = Table(td, colWidths=[2.2*cm,3.5*cm,2.5*cm,3.8*cm,2.3*cm,2.5*cm,2.3*cm,1.8*cm], repeatRows=1)
        ts = [('BACKGROUND',(0,0),(-1,0),colors.HexColor('#1F4E79')),
              ('TEXTCOLOR', (0,0),(-1,0),colors.white),
              ('FONTNAME',  (0,0),(-1,-1),BASE_FONT),
              ('FONTSIZE',  (0,0),(-1,0),9),('FONTSIZE',(0,1),(-1,-1),8),
              ('ALIGN',     (0,0),(-1,-1),'RIGHT'),
              ('GRID',      (0,0),(-1,-1),0.3,colors.HexColor('#CCCCCC')),
              ('TOPPADDING',(0,0),(-1,-1),4),('BOTTOMPADDING',(0,0),(-1,-1),4)]
        for i,(_, row) in enumerate(agent_result.iterrows(),1):
            c = colors.HexColor('#FFCCCC') if row['עלייה %']>30 else colors.HexColor('#EBF3FB')
            ts.append(('BACKGROUND',(0,i),(-1,i),c))
        mt.setStyle(TableStyle(ts))
        story.append(mt)

    # ══════════════════════════════════════════════
    # עמוד 2 — פוליסות שהוסרו + פוליסות חדשות
    # ══════════════════════════════════════════════
    story.append(PageBreak())
    story += page_header('פוליסות שהוסרו ופוליסות חדשות')

    # פוליסות שנסגרו
    story.append(Paragraph(rh(f'פוליסות שנסגרו ({len(g)})'), sec_s))
    if len(g) > 0:
        gone_h = [rh('ת.ז'), rh('שם לקוח'), rh("מס' פוליסה"), rh('יצרן'), rh('פרמיה אחרונה')]
        gone_d = [gone_h]
        for _, row in g.iterrows():
            gone_d.append([
                rh(str(row.get(COL_ID,''))),
                rh(str(row.get(COL_FNAME,'')) + ' ' + str(row.get(COL_LNAME,''))),
                rh(str(row.get(COL_POLICY,''))),
                rh(str(row.get(COL_MFG,''))),
                f"₪{row.get(COL_PREMIUM,0):,.0f}" if pd.notna(row.get(COL_PREMIUM)) else '—',
            ])
        gt = Table(gone_d, colWidths=[2.5*cm,4.5*cm,3.0*cm,4.0*cm,3.0*cm], repeatRows=1)
        gt.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#7F7F7F')),
            ('TEXTCOLOR', (0,0),(-1,0),colors.white),
            ('FONTNAME',  (0,0),(-1,-1),BASE_FONT),
            ('FONTSIZE',  (0,0),(-1,0),9),('FONTSIZE',(0,1),(-1,-1),8),
            ('ALIGN',     (0,0),(-1,-1),'RIGHT'),
            ('GRID',      (0,0),(-1,-1),0.3,colors.HexColor('#CCCCCC')),
            ('TOPPADDING',(0,0),(-1,-1),4),('BOTTOMPADDING',(0,0),(-1,-1),4),
            ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,colors.HexColor('#F2F2F2')]),
        ]))
        story.append(gt)
    else:
        story.append(Paragraph(rh('לא נמצאו פוליסות שנסגרו'), sub_s))

    story.append(Spacer(1, 0.7*cm))

    # פוליסות חדשות
    story.append(Paragraph(rh(f'פוליסות חדשות ({len(n)})'), sec_s))
    if len(n) > 0:
        new_h = [rh('ת.ז'), rh('שם לקוח'), rh("מס' פוליסה"), rh('יצרן'), rh('פרמיה')]
        new_d = [new_h]
        for _, row in n.iterrows():
            new_d.append([
                rh(str(row.get(COL_ID,''))),
                rh(str(row.get(COL_FNAME,'')) + ' ' + str(row.get(COL_LNAME,''))),
                rh(str(row.get(COL_POLICY,''))),
                rh(str(row.get(COL_MFG,''))),
                f"₪{row.get(COL_PREMIUM,0):,.0f}" if pd.notna(row.get(COL_PREMIUM)) else '—',
            ])
        nt = Table(new_d, colWidths=[2.5*cm,4.5*cm,3.0*cm,4.0*cm,3.0*cm], repeatRows=1)
        nt.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#375623')),
            ('TEXTCOLOR', (0,0),(-1,0),colors.white),
            ('FONTNAME',  (0,0),(-1,-1),BASE_FONT),
            ('FONTSIZE',  (0,0),(-1,0),9),('FONTSIZE',(0,1),(-1,-1),8),
            ('ALIGN',     (0,0),(-1,-1),'RIGHT'),
            ('GRID',      (0,0),(-1,-1),0.3,colors.HexColor('#CCCCCC')),
            ('TOPPADDING',(0,0),(-1,-1),4),('BOTTOMPADDING',(0,0),(-1,-1),4),
            ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,colors.HexColor('#EBF1DE')]),
        ]))
        story.append(nt)
    else:
        story.append(Paragraph(rh('לא נמצאו פוליסות חדשות'), sub_s))

    # ══════════════════════════════════════════════
    # עמוד 3 — חריגות דמי ניהול
    # ══════════════════════════════════════════════
    fe = fee_exceptions[fee_exceptions[COL_AGENT] == agent] if (fee_exceptions is not None and agent) else fee_exceptions
    if fe is not None and len(fe) > 0:
        try:
            story.append(PageBreak())
            story += page_header('חריגות דמי ניהול — מוצרי חיסכון')

            story.append(Paragraph(rh(f'מוצרים עם חריגה בדמי ניהול ({len(fe)})'), sec_s))
            fh = [rh('ת.ז'), rh('שם לקוח'), rh('סוג מוצר'), rh('מוצר'),
                  rh('צבירה'), rh('צבירה כוללת'), rh('דמי ניהול'), rh('סף מקסימלי'), rh('סיבת חריגה')]
            fd = [fh]
            for _, row in fe.iterrows():
                raw_fee    = row.get('דמי ניהול מצבירה', 0)
                raw_thresh = row.get('סף מקסימלי', 0)
                fee    = float(raw_fee)    if pd.notna(raw_fee)    else 0.0
                thresh = float(raw_thresh) if pd.notna(raw_thresh) else 0.0
                fd.append([
                    rh(str(row.get(COL_ID,''))),
                    rh(str(row.get('שם לקוח',''))),
                    rh(str(row.get('סוג מוצר',''))),
                    rh(str(row.get('מוצר',''))),
                    f"₪{row.get('צבירה',0):,.0f}",
                    f"₪{row.get('צבירה כוללת',0):,.0f}",
                    f"{fee*100:.3f}%",
                    f"{thresh*100:.2f}%",
                    rh(str(row.get('סיבת חריגה','')))
                ])
            ft = Table(fd, colWidths=[1.8*cm,2.8*cm,2.4*cm,2.8*cm,1.8*cm,1.9*cm,1.7*cm,1.7*cm,3.6*cm], repeatRows=1)
            fts = [
                ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#FF6600')),
                ('TEXTCOLOR', (0,0),(-1,0),colors.white),
                ('FONTNAME',  (0,0),(-1,-1),BASE_FONT),
                ('FONTSIZE',  (0,0),(-1,0),8),('FONTSIZE',(0,1),(-1,-1),7),
                ('ALIGN',     (0,0),(-1,-1),'RIGHT'),
                ('GRID',      (0,0),(-1,-1),0.3,colors.HexColor('#CCCCCC')),
                ('TOPPADDING',(0,0),(-1,-1),3),('BOTTOMPADDING',(0,0),(-1,-1),3),
            ]
            # צבע לפי קבוצת סף — כל רמת צבירה בצבע ייחודי
            def row_color(total):
                if total > 1_000_000: return colors.HexColor('#D6EAF8')  # תכלת — מעל 1M
                if total > 500_000:   return colors.HexColor('#FEF9E7')  # צהוב — מעל 500K
                if total > 250_000:   return colors.HexColor('#E9F7EF')  # ירוק — מעל 250K
                return colors.HexColor('#FDEDEC')                         # ורוד — מתחת 250K
            for i, (_, row) in enumerate(fe.iterrows(), 1):
                fts.append(('BACKGROUND', (0,i), (-1,i), row_color(row.get('צבירה כוללת', 0))))
            ft.setStyle(TableStyle(fts))
            story.append(ft)
        except Exception as e:
            story.append(Paragraph(rh(f'שגיאה בטעינת טבלת חריגות: {e}'), sub_s))

    # ══════════════════════════════════════════════
    # עמוד 4 — לקוחות זהב
    # ══════════════════════════════════════════════
    gc = gold_customers[gold_customers[COL_AGENT] == agent] if (gold_customers is not None and agent) else gold_customers
    if gc is not None and len(gc) > 0:
        try:
            story.append(PageBreak())
            story += page_header('לקוחות זהב — צבירה מעל ₪1,000,000')
            story.append(Paragraph(rh(f'לקוחות עם צבירה כוללת מעל מיליון ש״ח ({len(gc)})'), sec_s))

            # כותרות קצרות לעמודות
            TYPE_SHORT = {'קרן השתלמות': 'השתלמות', 'קופת גמל': 'קופת גמל', 'קופת גמל להשקעה': 'גמל להשקעה'}
            gold_type_cols = [c for c in ['קרן השתלמות', 'קופת גמל', 'קופת גמל להשקעה'] if c in gc.columns]
            has_fee = 'דמי ניהול ממוצע' in gc.columns
            gh = [rh('ת.ז'), rh('שם לקוח'), rh('מת"ל'), rh('צבירה כוללת')]
            gh += [rh(TYPE_SHORT.get(c, c)) for c in gold_type_cols]
            if has_fee:
                gh.append(rh('דמי ניהול'))
            gd = [gh]
            for _, row in gc.iterrows():
                gr = [
                    rh(str(row.get(COL_ID,''))),
                    rh(str(row.get('שם לקוח',''))),
                    rh(str(row.get(COL_AGENT,''))),
                    f"₪{row.get('צבירה כוללת',0):,.0f}",
                ]
                for c in gold_type_cols:
                    v = row.get(c, 0)
                    gr.append(f"₪{v:,.0f}" if v > 0 else '—')
                if has_fee:
                    fee_v = row.get('דמי ניהול ממוצע')
                    gr.append(f"{fee_v*100:.3f}%" if pd.notna(fee_v) else '—')
                gd.append(gr)

            ncols = len(gh)
            col_ws = [1.8*cm, 3.2*cm, 2.3*cm, 2.2*cm] + [2.0*cm] * (ncols - 4)

            gt = Table(gd, colWidths=col_ws, repeatRows=1)
            gts = [
                ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#B8860B')),
                ('TEXTCOLOR', (0,0),(-1,0),colors.white),
                ('FONTNAME',  (0,0),(-1,-1),BASE_FONT),
                ('FONTSIZE',  (0,0),(-1,0),9),('FONTSIZE',(0,1),(-1,-1),8),
                ('ALIGN',     (0,0),(-1,-1),'RIGHT'),
                ('GRID',      (0,0),(-1,-1),0.3,colors.HexColor('#CCCCCC')),
                ('TOPPADDING',(0,0),(-1,-1),4),('BOTTOMPADDING',(0,0),(-1,-1),4),
            ]
            for i in range(1, len(gd)):
                bg = colors.HexColor('#FFFACD') if i % 2 == 1 else colors.HexColor('#FFF8DC')
                gts.append(('BACKGROUND',(0,i),(-1,i),bg))
            gt.setStyle(TableStyle(gts))
            story.append(gt)
        except Exception as e:
            story.append(Paragraph(rh(f'שגיאה בטעינת טבלת לקוחות זהב: {e}'), sub_s))

    # ══════════════════════════════════════════════
    # עמוד 5 — חריגות פנסיה
    # ══════════════════════════════════════════════
    pd1 = pension_deposit_exc[pension_deposit_exc[COL_AGENT]==agent] if (pension_deposit_exc is not None and agent) else pension_deposit_exc
    pa1 = pension_acc_exc[pension_acc_exc[COL_AGENT]==agent]         if (pension_acc_exc     is not None and agent) else pension_acc_exc
    has_p1 = pd1 is not None and len(pd1) > 0
    has_p2 = pa1 is not None and len(pa1) > 0

    if has_p1 or has_p2:
        try:
            story.append(PageBreak())
            story += page_header('חריגות דמי ניהול — קרן פנסיה')

            def pension_row_color(total):
                if total > 1_000_000: return colors.HexColor('#D6EAF8')  # תכלת — מעל 1M
                if total > 500_000:   return colors.HexColor('#FEF9E7')  # צהוב — מעל 500K
                return colors.HexColor('#E9F7EF')                         # ירוק — מעל 200K

            def pension_table(df_p, title, hdr_color, cols_def, color_by_savings=False):
                """cols_def = [(header, col_key, fmt_fn)]"""
                story.append(Paragraph(rh(title), sec_s))
                # סגנונות לכותרות ולתאי גוף (עם גלישת טקסט — ללא RTL כפול)
                hdr_cell_s = ParagraphStyle('phdr', fontName=BASE_FONT, fontSize=8,
                                            textColor=colors.white, alignment=2, leading=10)
                body_cell_s = ParagraphStyle('pbody', fontName=BASE_FONT, fontSize=7,
                                             textColor=colors.black, alignment=2, leading=9)
                def _ph(txt): return Paragraph(rh(str(txt)), hdr_cell_s)
                def _pb(txt): return Paragraph(rh(str(txt)), body_cell_s)

                th = [_ph(h) for h,_,_ in cols_def]
                td = [th]
                rows_list = list(df_p.iterrows())
                for _, row in rows_list:
                    td.append([_pb(fmt(row.get(k,''))) for _,k,fmt in cols_def])
                # רוחבי עמודות: ת.ז | שם לקוח | מת"ל | סוג מוצר | חברה | צבירה | צבירה כוללת | דמי צבירה | סיבת חריגה
                col_ws = [1.8*cm, 2.8*cm, 3.2*cm, 1.8*cm, 1.8*cm, 1.8*cm, 1.8*cm, 1.5*cm, 1.5*cm]
                t = Table(td, colWidths=col_ws[:len(cols_def)], repeatRows=1)
                ts = [
                    ('BACKGROUND',(0,0),(-1,0),colors.HexColor(hdr_color)),
                    ('FONTNAME',  (0,0),(-1,-1),BASE_FONT),
                    ('ALIGN',     (0,0),(-1,-1),'RIGHT'),
                    ('VALIGN',    (0,0),(-1,-1),'TOP'),
                    ('GRID',      (0,0),(-1,-1),0.3,colors.HexColor('#CCCCCC')),
                    ('TOPPADDING',(0,0),(-1,-1),3),('BOTTOMPADDING',(0,0),(-1,-1),3),
                ]
                for i, (_, row) in enumerate(rows_list, 1):
                    if color_by_savings:
                        bg = pension_row_color(row.get('צבירה כוללת', 0))
                    else:
                        bg = colors.HexColor('#F4ECF7') if i%2==1 else colors.HexColor('#EBD5F7')
                    ts.append(('BACKGROUND',(0,i),(-1,i),bg))
                t.setStyle(TableStyle(ts))
                story.append(t)
                story.append(Spacer(1, 0.5*cm))

            if has_p1:
                pension_table(pd1, f'טבלה 1 — חריגות דמי ניהול מהפקדה >2% ({len(pd1)})', '#6A0DAD', [
                    ('ת.ז',        COL_ID,                  lambda v: rh(str(v))),
                    ('שם לקוח',    'שם לקוח',               lambda v: rh(str(v))),
                    ('מת"ל',       COL_AGENT,                lambda v: rh(str(v))),
                    ('סוג מוצר',   'סוג מוצר',              lambda v: rh(str(v))),
                    ('חברה',       'יצרן',                   lambda v: rh(str(v))),
                    ('צבירה',      'צבירה',                  lambda v: f"₪{v:,.0f}" if pd.notna(v) and v else '—'),
                    ('דמי הפקדה',  'דמי ניהול מהפקדה',      lambda v: f"{float(v)*100:.3f}%" if pd.notna(v) and v!='' else '—'),
                ])

            if has_p2:
                pension_table(pa1, f'טבלה 2 — חריגות דמי ניהול מצבירה ({len(pa1)})', '#154360', [
                    ('ת.ז',        COL_ID,                  lambda v: rh(str(v))),
                    ('שם לקוח',    'שם לקוח',               lambda v: rh(str(v))),
                    ('מת"ל',       COL_AGENT,                lambda v: rh(str(v))),
                    ('סוג מוצר',   'סוג מוצר',              lambda v: rh(str(v))),
                    ('חברה',       'יצרן',                   lambda v: rh(str(v))),
                    ('צבירה',      'צבירה',                  lambda v: f"₪{v:,.0f}" if pd.notna(v) and v else '—'),
                    ('צבירה כוללת','צבירה כוללת',            lambda v: f"₪{v:,.0f}" if pd.notna(v) and v else '—'),
                    ('דמי צבירה',  'דמי ניהול מצבירה',      lambda v: f"{float(v)*100:.3f}%" if pd.notna(v) and v!='' else '—'),
                    ('סיבת חריגה', 'סיבת חריגה',            lambda v: rh(str(v))),
                ], color_by_savings=True)
        except Exception as e:
            story.append(Paragraph(rh(f'שגיאה בטעינת טבלת פנסיה: {e}'), sub_s))

    doc.build(story)
    return buf.getvalue()

# ══════════════════════════════════════════════════════════════════════════════
# PASSWORD PROTECTION
# ══════════════════════════════════════════════════════════════════════════════
def check_password():
    correct = st.secrets.get("APP_PASSWORD", "surense2025")
    if "authenticated" not in st.session_state:
        st.session_state.authenticated = False
    if st.session_state.authenticated:
        return True
    st.markdown("""
    <div style="max-width:360px;margin:80px auto;text-align:right;direction:rtl">
        <div style="font-size:2rem;text-align:center;margin-bottom:8px">📊</div>
        <h2 style="text-align:center;color:#1F4E79;margin-bottom:24px">מערכת דוחות פרמיה</h2>
    </div>
    """, unsafe_allow_html=True)
    col_l, col_m, col_r = st.columns([1,2,1])
    with col_m:
        pwd = st.text_input("🔒 סיסמה", type="password", placeholder="הכנס סיסמה...")
        if st.button("כניסה", use_container_width=True):
            if pwd == correct:
                st.session_state.authenticated = True
                st.rerun()
            else:
                st.error("סיסמה שגויה — נסה שוב")
    st.stop()

check_password()

# ══════════════════════════════════════════════════════════════════════════════
# UI
# ══════════════════════════════════════════════════════════════════════════════
st.title("📊 מערכת דוחות פרמיה")
st.markdown("העלה את שני דוחות ה-Excel מ-Surense וקבל דוחות מוכנים להורדה.")
st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)

col1, col2 = st.columns(2)
with col1:
    st.markdown("**📁 דוח החודש הקודם**")
    f1 = st.file_uploader("", type=['xlsx'], key='f1', label_visibility='collapsed')
with col2:
    st.markdown("**📁 דוח החודש הנוכחי**")
    f2 = st.file_uploader("", type=['xlsx'], key='f2', label_visibility='collapsed')

if f1 and f2:
    st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)
    with st.spinner("מנתח נתונים..."):
        try:
            f1_bytes = f1.read()
            f2_bytes = f2.read()
            merged, result, gone_df, new_df, df1d, df2d = analyze(f1_bytes, f2_bytes)
            fee_exceptions, fee_breakdown = analyze_management_fees(f2_bytes)
            pension_deposit_exc, pension_acc_exc = analyze_pension_fees(f2_bytes)
            gold_customers = analyze_gold_customers(f2_bytes)
            agents = sorted(merged['מת"ל'].dropna().unique().tolist())
            month_label = f'{f1.name[:10]} ← {f2.name[:10]}'
        except Exception as e:
            st.error(f"שגיאה בקריאת הקבצים: {e}")
            st.stop()



    # ── Summary metrics ──
    st.subheader("📈 סיכום")
    c1, c2, c3, c4 = st.columns(4)
    c1.markdown(f'<div class="metric-box"><div class="val">{len(merged)}</div><div class="lbl">פוליסות שהושוו</div></div>', unsafe_allow_html=True)
    c2.markdown(f'<div class="metric-box"><div class="val red">{len(result)}</div><div class="lbl">עלייה >15%</div></div>', unsafe_allow_html=True)
    c3.markdown(f'<div class="metric-box"><div class="val">{len(gone_df)}</div><div class="lbl">פוליסות שנסגרו</div></div>', unsafe_allow_html=True)
    c4.markdown(f'<div class="metric-box"><div class="val green">{len(new_df)}</div><div class="lbl">פוליסות חדשות</div></div>', unsafe_allow_html=True)

    if len(result) > 0:
        st.markdown(f"""
        <br>
        <div style="background:#FFF3CD;border-radius:10px;padding:16px 20px;border:2px solid #FF9500;text-align:right;direction:rtl">
            <div style="font-size:1.05rem;font-weight:bold;color:#7A4F00;margin-bottom:10px">⚠️ סיכום עליות פרמיה</div>
            <div style="display:flex;justify-content:flex-end;gap:32px;flex-wrap:wrap">
                <div style="text-align:center">
                    <div style="font-size:1.4rem;font-weight:bold;color:#CC3300">₪{result['עלייה ₪'].sum():,.0f}</div>
                    <div style="font-size:0.8rem;color:#7A4F00;margin-top:2px">סך עלייה חודשית</div>
                </div>
                <div style="text-align:center">
                    <div style="font-size:1.4rem;font-weight:bold;color:#CC3300">{result['עלייה %'].mean():.1f}%</div>
                    <div style="font-size:0.8rem;color:#7A4F00;margin-top:2px">ממוצע עלייה</div>
                </div>
                <div style="text-align:center">
                    <div style="font-size:1.4rem;font-weight:bold;color:#CC3300">{result['עלייה %'].max():.1f}%</div>
                    <div style="font-size:0.8rem;color:#7A4F00;margin-top:2px">עלייה מקסימלית</div>
                </div>
                <div style="text-align:center">
                    <div style="font-size:1.4rem;font-weight:bold;color:#1F4E79">{len(result)}</div>
                    <div style="font-size:0.8rem;color:#7A4F00;margin-top:2px">פוליסות עם עלייה</div>
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)

    st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)

    # ── Preview table ──
    if len(result) > 0:
        st.subheader("🔍 פוליסות עם עלייה >15%")
        preview = result[['שם לקוח','מת"ל','יצרן','פרמיה קודמת','פרמיה נוכחית','עלייה ₪','עלייה %']].copy()
        preview['עלייה %'] = preview['עלייה %'].map(lambda x: f"{x:.1f}%")
        preview['פרמיה קודמת'] = preview['פרמיה קודמת'].map(lambda x: f"₪{x:,.0f}")
        preview['פרמיה נוכחית'] = preview['פרמיה נוכחית'].map(lambda x: f"₪{x:,.0f}")
        preview['עלייה ₪'] = preview['עלייה ₪'].map(lambda x: f"₪{x:,.0f}")
        st.dataframe(preview, use_container_width=True, hide_index=True)

    # ── Fee exceptions preview ──
    if len(fee_exceptions) > 0:
        st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)
        st.subheader(f"🚨 חריגות דמי ניהול — {len(fee_exceptions)} מוצרים")
        st.markdown(f"""
        <div style="background:#7A3000;border-radius:10px;padding:12px 16px;border:2px solid #FF6600;text-align:right;direction:rtl;margin-bottom:12px;color:#FFD6B0;">
        🔶 נמצאו לקוחות שמשלמים דמי ניהול מצבירה מעל הסף המותר לפי גובה הצבירה שלהם
        </div>
        """, unsafe_allow_html=True)
        cols = ['שם לקוח', COL_AGENT, 'סוג מוצר']
        if 'מוצר' in fee_exceptions.columns:
            cols.append('מוצר')
        cols += ['צבירה', 'צבירה כוללת', 'דמי ניהול מצבירה', 'סף מקסימלי', 'סיבת חריגה']
        fee_preview = fee_exceptions[[c for c in cols if c in fee_exceptions.columns]].copy()
        if 'צבירה' in fee_preview.columns:
            fee_preview['צבירה']           = fee_preview['צבירה'].map(lambda x: f"₪{x:,.0f}")
        fee_preview['צבירה כוללת']        = fee_preview['צבירה כוללת'].map(lambda x: f"₪{x:,.0f}")
        fee_preview['דמי ניהול מצבירה']   = fee_preview['דמי ניהול מצבירה'].map(lambda x: f"{x*100:.3f}%")
        fee_preview['סף מקסימלי']          = fee_preview['סף מקסימלי'].map(lambda x: f"{x*100:.2f}%")
        st.dataframe(fee_preview, use_container_width=True, hide_index=True)

    # ── Gold customers ──
    # ── Pension fee exceptions ──
    if len(pension_deposit_exc) > 0 or len(pension_acc_exc) > 0:
        st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)
        st.subheader("🏦 חריגות דמי ניהול — קרן פנסיה")

        if len(pension_deposit_exc) > 0:
            st.markdown(f"**טבלה 1 — חריגות דמי ניהול מהפקדה ({len(pension_deposit_exc)} מוצרים)**")
            st.markdown("""<div style="background:#4A1A6B;border-radius:8px;padding:10px 14px;border:2px solid #9B59B6;text-align:right;direction:rtl;margin-bottom:10px;color:#E8DAEF;">
            🔵 לקוחות שמשלמים יותר מ-2% דמי ניהול מהפקדה</div>""", unsafe_allow_html=True)
            p1 = pension_deposit_exc[[c for c in ['שם לקוח', COL_AGENT, 'סוג מוצר', 'יצרן', 'מוצר', 'צבירה', 'דמי ניהול מהפקדה'] if c in pension_deposit_exc.columns]].copy()
            if 'צבירה' in p1.columns: p1['צבירה'] = p1['צבירה'].map(lambda x: f"₪{x:,.0f}")
            if 'דמי ניהול מהפקדה' in p1.columns: p1['דמי ניהול מהפקדה'] = p1['דמי ניהול מהפקדה'].map(lambda x: f"{x*100:.3f}%")
            st.dataframe(p1, use_container_width=True, hide_index=True)

        if len(pension_acc_exc) > 0:
            st.markdown(f"**טבלה 2 — חריגות דמי ניהול מצבירה ({len(pension_acc_exc)} מוצרים)**")
            st.markdown("""<div style="background:#154360;border-radius:8px;padding:10px 14px;border:2px solid #2E86C1;text-align:right;direction:rtl;margin-bottom:10px;color:#D6EAF8;">
            🔷 לקוחות שמשלמים דמי ניהול מצבירה מעל הסף לפי גובה הצבירה</div>""", unsafe_allow_html=True)
            p2 = pension_acc_exc[[c for c in ['שם לקוח', COL_AGENT, 'סוג מוצר', 'יצרן', 'מוצר', 'צבירה', 'צבירה כוללת', 'דמי ניהול מצבירה', 'סף מקסימלי', 'סיבת חריגה'] if c in pension_acc_exc.columns]].copy()
            if 'צבירה' in p2.columns: p2['צבירה'] = p2['צבירה'].map(lambda x: f"₪{x:,.0f}")
            if 'צבירה כוללת' in p2.columns: p2['צבירה כוללת'] = p2['צבירה כוללת'].map(lambda x: f"₪{x:,.0f}")
            if 'דמי ניהול מצבירה' in p2.columns: p2['דמי ניהול מצבירה'] = p2['דמי ניהול מצבירה'].map(lambda x: f"{x*100:.3f}%")
            if 'סף מקסימלי' in p2.columns: p2['סף מקסימלי'] = p2['סף מקסימלי'].map(lambda x: f"{x*100:.3f}%")
            st.dataframe(p2, use_container_width=True, hide_index=True)

    if len(gold_customers) > 0:
        st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)
        st.subheader(f"🏆 לקוחות זהב — {len(gold_customers)} לקוחות עם צבירה מעל ₪1M")
        st.markdown(f"""
        <div style="background:#7A6000;border-radius:10px;padding:12px 16px;border:2px solid #FFD700;text-align:right;direction:rtl;margin-bottom:12px;color:#FFF8DC;">
        ⭐ לקוחות עם צבירה כוללת מעל ₪1,000,000 במוצרי חיסכון (קרן השתלמות, קופת גמל, גמל להשקעה)
        </div>
        """, unsafe_allow_html=True)
        gold_cols = ['שם לקוח', COL_AGENT, 'צבירה כוללת']
        for c in ['קרן השתלמות', 'קופת גמל', 'קופת גמל להשקעה', 'דמי ניהול ממוצע']:
            if c in gold_customers.columns:
                gold_cols.append(c)
        gold_preview = gold_customers[[c for c in gold_cols if c in gold_customers.columns]].copy()
        gold_preview['צבירה כוללת'] = gold_preview['צבירה כוללת'].map(lambda x: f"₪{x:,.0f}")
        for c in ['קרן השתלמות', 'קופת גמל', 'קופת גמל להשקעה']:
            if c in gold_preview.columns:
                gold_preview[c] = gold_preview[c].map(lambda x: f"₪{x:,.0f}" if x > 0 else '—')
        if 'דמי ניהול ממוצע' in gold_preview.columns:
            gold_preview['דמי ניהול ממוצע'] = gold_preview['דמי ניהול ממוצע'].map(lambda x: f"{x*100:.3f}%" if pd.notna(x) else '—')
        st.dataframe(gold_preview, use_container_width=True, hide_index=True)

    st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)

    # ── Downloads ──
    st.subheader("📥 הורדת דוחות")

    # Combined
    with st.spinner("בונה דוח כולל..."):
        xl_all  = build_excel(merged, result, gone_df, new_df, fee_exceptions=fee_exceptions, gold_customers=gold_customers, pension_deposit_exc=pension_deposit_exc, pension_acc_exc=pension_acc_exc)
        pdf_all = build_pdf(merged, result, gone_df, new_df, month_label, fee_exceptions=fee_exceptions, gold_customers=gold_customers, pension_deposit_exc=pension_deposit_exc, pension_acc_exc=pension_acc_exc)

    st.markdown("**דוח כולל — כל הסוכנים**")
    ca, cb = st.columns(2)
    with ca:
        st.download_button("⬇️ הורד Excel", xl_all,
                           file_name=f"דוח_פרמיה_{datetime.now().strftime('%m_%Y')}.xlsx",
                           mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    with cb:
        st.download_button("⬇️ הורד PDF", pdf_all,
                           file_name=f"דוח_פרמיה_{datetime.now().strftime('%m_%Y')}.pdf",
                           mime='application/pdf')

    # Per agent
    st.markdown("**דוחות לפי סוכן מטפל**")
    for agent in agents:
        with st.spinner(f"בונה דוח עבור {agent}..."):
            xl_a  = build_excel(merged, result, gone_df, new_df, agent=agent)
            pdf_a = build_pdf(merged, result, gone_df, new_df, month_label, agent=agent, fee_exceptions=fee_exceptions, gold_customers=gold_customers, pension_deposit_exc=pension_deposit_exc, pension_acc_exc=pension_acc_exc)
        safe = agent.replace(' ','_')
        st.markdown(f"**{agent}**")
        da, db = st.columns(2)
        with da:
            st.download_button(f"⬇️ Excel — {agent}", xl_a,
                               file_name=f"דוח_פרמיה_{safe}_{datetime.now().strftime('%m_%Y')}.xlsx",
                               mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                               key=f'xl_{safe}')
        with db:
            st.download_button(f"⬇️ PDF — {agent}", pdf_a,
                               file_name=f"דוח_פרמיה_{safe}_{datetime.now().strftime('%m_%Y')}.pdf",
                               mime='application/pdf',
                               key=f'pdf_{safe}')

else:
    st.info("⬆️ העלה את שני קבצי ה-Excel כדי להתחיל")
